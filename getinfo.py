import re
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
import requests
from bs4 import BeautifulSoup
import json

wb = load_workbook(filename='output.xlsx')
ws = wb.active
urls = []
def readfile(filename):
	urlfile = open(filename, "r")
	for x in urlfile:
		urls.append(x)
	urlfile.close()
	print(len(urls))
readfile('urls.txt')

count = 0

def formatString(stri):
	return stri.replace("\t", '').replace("\n", ' ').replace("  ", '')

for url in urls:

	count += 1
	# if count % 2 > 0:
	# 	continue

	print(url)

	curl = url.replace("\n", "")
	res = requests.get(curl)
	soup = BeautifulSoup(res.text, "html.parser")
	titleRaw = soup.select('div.pdp-content h3')[0].getText()
	title = formatString(titleRaw)
	texture = soup.select('span#pdpdata-texture')[0].getText()
	design = soup.select('span#pdpdata-design')[0].getText()
	fiber_content = soup.select('span#pdpdata-fiber_content')[0].getText()
	construction = soup.select('span#pdpdata-construction')[0].getText()
	country_of_origin = soup.select('span#pdpdata-country_of_origin')[0].getText()
	repeat_width_length = soup.select('span#pdpdata-repeat_width_length')[0].getText()
	roll_width = soup.select('span#pdpdata-roll_width')[0].getText()
	repeat_width = soup.select('span#pdpdata-repeat_width')[0].getText()
	repeat_length = soup.select('span#pdpdata-repeat_length')[0].getText()
	collection = soup.select('div.detail-value')[0].getText()

	variants = soup.select('div.variant-item a')
	for variant in variants:
		sku = variant['data-api-sku']
		color = variant['data-color-name']
		api = variant['data-api-feature_image']
		
		print(title, sku)
		
		imageUrl = api.replace("?$med_thumb$", "_SET?req=set,json,UTF-8")
		response = requests.get(imageUrl)
		if response.status_code == 200:
			imageData = response.content
			start_index = imageData.index(b'(') + 1
			end_index = imageData.rindex(b')') - 3
			json_data = imageData[start_index:end_index]
			parsed_data = json.loads(json_data)
			props = ["CARPET", "PRESTIGEMILLS", sku, title, '', collection, color, texture, design, fiber_content, construction, country_of_origin, repeat_width_length, repeat_width, repeat_length, roll_width]
			try:
				images = parsed_data['set']['item']
				if type(images) != list:
					images = [images]
				for image in images:
					width = "642"
					height = int(int(image['dy'])*642/int(image['dx']))
					props.append("https://s7d2.scene7.com/is/image/"+image['i']['n']+"?wid="+width+"&hei="+str(height))
				ws.append(props)
			except Exception as e:
				print('error=>', url)
	if count >= 20:
		wb.save('output-new.xlsx')
		print('saved', count)
		count = 0
	time.sleep(0.2)
wb.save('output-new.xlsx')